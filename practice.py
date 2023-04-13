import pandas as pd
import streamlit as st
import pandas as pd
import plotly.express as px

st.set_page_config(page_title='Doppio Zero')
st.title('Excel Plotter')
st.subheader('Feed me with excel files')

uploaded_file = st.file_uploader('Choose an XLSX file ',type = 'xlsx')
if uploaded_file:
    st.markdown('---')
    df = pd.read_excel(uploaded_file, engine='openpyxl')
    st.dataframe(df)
    groupby_column = st.selectbox(
        'What would you like to analyse?',
        ('Ship Mode', 'Segment', 'Category', 'Sub-Category'),
    )

    #---Dataframe
    output_columns = ['Sales', 'Profit']
    df_grouped = df.groupby(by=[groupby_column], as_index=False)[output_columns].sum()
    st.dataframe(df_grouped)

    #copy and paste project
    from openpyxl import load_workbook

wb = load_workbook('kpi_day.xlsx')
sheet1 = wb['Report']
sheet2 = wb['Target']

# for i in range(10, 15):
#     for j in range(2, 3):
#         sheet2.cell(row=5, column=6).value = sheet1.cell(row=i, column=j).value

for i in range(10, 20):
    for j in range(3, 4):
        sheet2.cell(row=i-6, column=j).value = sheet1.cell(row=i, column=j).value

wb.save('kpi_master.xlsx')





